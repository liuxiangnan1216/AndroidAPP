#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2018/12/13 15:54
# @Author  : LiuXiangNan
# @Site    : 
# @File    : LogAnalysis.py
import datetime
import difflib
import sys
import threading
from tkinter import filedialog
import tkinter.messagebox
import tkinter as tk

import openpyxl
import xlsxwriter as xlsxwriter
from openpyxl import styles
from PyQt5 import QtWidgets
root = tk.Tk()
root.withdraw()
from MonkeyLogParseUI import Ui_MainWindow
import configparser
from AnalysisLog2 import *
config = configparser.ConfigParser()
config.read('setting.ini', encoding='utf-8')


class MainWindow(QtWidgets.QMainWindow, Ui_MainWindow):
    def __init__(self, parent=None):
        super(MainWindow, self).__init__(parent)
        self.setupUi(self)

        ################################################################################################################
        self.toolButton_path.clicked.connect(self.select_path)
        self.pushButton_start.clicked.connect(self.start_parse_thread)
        self.pushButton_about.clicked.connect(about)
        self.pushButton_help.clicked.connect(softHelp)

        ################################################################################################################

    def start_parse_thread(self):
        project_name = self.lineEdit_name.text()
        log_path = self.lineEdit_path.text()

        java = subprocess.Popen('java', shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE).communicate()

        print('java====', str(java[1]))

        if re.findall(r'java \[-options\] class \[args...\]', str(java[1])):
            print('9999999999999999999999')
        if not project_name:
            tk.messagebox.showerror('ERROR', '请选择正确的项目名！')
            return
        if not log_path:
            tk.messagebox.showerror('ERROR', '请选择正确的文件路径！')
            return

        startThread = threading.Thread(target=self.start_parse, name='start_parse')
        startThread.start()


    def start_parse(self):
        # self.import_report()
        patteran_we = r'( W )|( E )'
        pattern_name = r'main_log'
        java_exception = config.get('EXCEPTION_MAP', 'EXCEPTION')
        exception_list = java_exception.split('|')
        log_path = self.lineEdit_path.text()
        project_name = self.lineEdit_name.text()
        dirsName = self.get_DevicesID_dirs(log_path)
        conn = sqlite3.connect('./MonkeyParseLog.db')
        cursor = conn.cursor()
        table_name = 'ORIGIN_Exception'
        cursor.execute('DROP TABLE IF EXISTS {0}'.format(table_name))
        cursor.execute('''CREATE TABLE {0} (
                                           db_id INT PRIMARY KEY,
                                           root STRING,
                                           row_num STRING,
                                           date STRING,
                                           time STRING,
                                           pid STRING,
                                           tid STRING,
                                           e_w STRING,
                                           tag STRING,
                                           Ex_info STRING
                                           )'''.format(table_name))

        db_id = 0
        insert_data = []
        for i in range(0, len(dirsName)):
            print('devices id =========', dirsName)
            path1 = log_path + '\\' + dirsName[i]
            files_log_list = self.parse_all_log(path1)
            for n in range(0, len(files_log_list)):
                if re.search(pattern_name, files_log_list[n]):
                    fp = open(files_log_list[n], 'rb')
                    # print('files_log_list==========', n, 'name==', files_log_list[n])
                    row_num = 0
                    while True:
                        line = fp.readline()
                        if not line:
                            break
                        else:
                            row_num = row_num + 1
                            for k in range(0, len(exception_list)):
                                if re.search(r'%s' % (exception_list[k]), str(line), re.I):
                                    if re.search(patteran_we, str(line)):
                                        print('files_log_list==========', n, '\nname==', files_log_list[n])
                                        split_result = self.split_list(line)
                                        db_id = db_id + 1
                                        insert_db = [db_id, files_log_list[n], row_num, split_result[0], split_result[1],
                                                     split_result[2], split_result[3], split_result[4], split_result[5],
                                                     split_result[6]]
                                        print('insert_db==============', insert_db)
                                        # print('n=====================', n)
                                        cursor.executemany('INSERT INTO ORIGIN_Exception VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)', [insert_db])
                    conn.commit()
                    fp.close()

        cursor.execute('SELECT * from {0}'.format(table_name))
        result1 = list(cursor.fetchall())

        cursor.execute('delete from {0} where db_id not in(select max(db_id) from {1} group by date,pid,tag,Ex_info)'
                       .format(table_name, table_name))
        cursor.execute('SELECT * from {0}'.format(table_name))
        result2 = list(cursor.fetchall())

        cursor.execute('select * from {0} order by tag'.format(table_name))
        result3 = cursor.fetchall()

        result_second = self.second_dup(list(result3))
        (xlsxname, orex, dupex) = self.import_report(project_name, list(result1), list(result2), list(result_second))
        conn.commit()
        print('orining exception end')

        ################################################################################################################
        main_analysislog(path_log=log_path)

        # 生成最终的报告
        self.import_report_end(xlsxname, orex, dupex)

    # def test(self):
    #     xlsxname = './report/lll_Monkey测试结果_201812210856.xlsx'
    #     orex = 10
    #     dupex = 20
    #     self.import_report_end(xlsxname, orex, dupex)


    def import_report_end(self, name, orex, dupex):
        anr_num = 0
        anrm_num = 0
        ex_num = 0
        exm_num = 0
        workbook = openpyxl.load_workbook(filename=name)
        # workbook.re
        wreport = workbook.create_sheet(title='测试结果信息及说明', index=0)
        wanr = workbook.create_sheet(title='ANR', index=3)
        wanrm = workbook.create_sheet(title='ANR_module', index=4)
        wex = workbook.create_sheet(title='Exception', index=5)
        wexm = workbook.create_sheet(title='Exception_module', index=6)

        fp_anr = open('./report/ANR.txt', 'r', encoding='utf-8')
        readLine_anr = fp_anr.readlines()
        for i in range(0, len(readLine_anr)):
            wanr.cell(i + 1, 1, readLine_anr[i]).value
            anr_num = i + 1
        fp_anr.close()

        fp_anr_mod = open('./report/ANR_module.txt', 'r', encoding='utf-8')
        readLine_anr_mod = fp_anr_mod.readlines()
        for j in range(0, len(readLine_anr_mod)):
            wanrm.cell(j + 1, 1, readLine_anr_mod[j]).value
            anrm_num = j + 1
        fp_anr_mod.close()

        fp_ex = open('./report/EXCEPTION.txt', 'r', encoding='utf-8')
        readLine_ex = fp_ex.readlines()
        for k in range(0, len(readLine_ex)):
            wex.cell(k + 1, 1, readLine_ex[k]).value
            ex_num = k + 1
        fp_ex.close()

        fp_ex_mod = open('./report/Exception_module.txt', 'r', encoding='utf-8')
        readLine_ex_mod = fp_ex_mod.readlines()
        for n in range(0, len(readLine_ex_mod)):
            wexm.cell(n + 1, 1, readLine_ex_mod[n]).value
            exm_num = n + 1
        fp_ex_mod.close()

        wreport.cell(1, 1, '类别')
        wreport.cell(1, 2, '数目')
        # wreport.row_dimensions[1].height = 100
        # wreport.row_dimensions[2].height = 100
        wreport.column_dimensions['A'].width = 40
        wreport.column_dimensions['B'].width = 40
        wreport.cell(1, 1).fill = styles.fills.GradientFill(stop=['008000', '008000'])
        wreport.cell(1, 2).fill = styles.fills.GradientFill(stop=['008000', '008000'])
        # wreport.cell(1, 2).alignment

        wreport.cell(2, 1, 'main log 发生 Exception 共：')
        wreport.cell(3, 1, '相似度 >80% 去重后共计：')
        wreport.cell(4, 1, '发生 ANR 共计：')
        wreport.cell(5, 1, '发生 ANR 的 Module 共计：')
        wreport.cell(6, 1, '发生重要 Exception 共计：')
        wreport.cell(7, 1, '发生 Exception Module共计：')

        wreport.cell(2, 2, orex).alignment = openpyxl.styles.Alignment(horizontal='left')
        wreport.cell(3, 2, dupex).alignment = openpyxl.styles.Alignment(horizontal='left')
        wreport.cell(4, 2, anr_num).alignment = openpyxl.styles.Alignment(horizontal='left')
        wreport.cell(5, 2, anrm_num).alignment = openpyxl.styles.Alignment(horizontal='left')
        wreport.cell(6, 2, ex_num).alignment = openpyxl.styles.Alignment(horizontal='left')
        wreport.cell(7, 2, exm_num).alignment = openpyxl.styles.Alignment(horizontal='left')




        workbook.save(filename=name)
        workbook.close()
        print('end')



    def write_buffer_txt(self, list_write):
        fpw = open('./buffer.txt', 'w', encoding='utf-8')
        fpw.write(list_write)
        fpw.write('\n')




    def import_report(self, name, msg1, msg2, msg3):
        list_msg1 = list(msg1)
        list_msg2 = list(msg2)
        list_msg3 = list(msg3)
        sheet_name = '{0}_Monkey测试结果'.format(name)
        print('sheet_name===================', sheet_name)
        print('dir is =========', Path.cwd())
        date = datetime.datetime.now().strftime('%Y%m%d%H%M')  # 当前时间
        workbookname = './report/{0}_{1}.xlsx'.format(sheet_name, date)
        workbook = xlsxwriter.Workbook(workbookname)
        # worksheet1 = workbook.add_worksheet('测试结果信息及说明')
        # worksheet2 = workbook.add_worksheet('原始异常信息')
        worksheet2 = workbook.add_worksheet('相似度100%去重后')
        worksheet3 = workbook.add_worksheet('相似度>80%去重后')
        bold_title = workbook.add_format({'bold': 1, 'bg_color': 'green', 'align': 'center', 'font_size': 14})
        title = ['ID', '路径', '行号', '日期', '时间', 'PID', 'TID', 'E/W', 'TAG', 'Exception info']

        worksheet2.set_column('A:H', 8)
        worksheet2.set_column('I:I', 30)
        worksheet2.set_column('J:J', 100)
        worksheet2.set_row(0, 20)
        worksheet3.set_column('A:H', 8)
        worksheet3.set_column('I:I', 30)
        worksheet3.set_column('J:J', 100)
        worksheet3.set_row(0, 20)
        # worksheet4.set_column('A:H', 8)
        # worksheet4.set_column('I:I', 30)
        # worksheet4.set_column('J:J', 100)
        # worksheet4.set_row(0, 20)

        # 写各个sheet表的标题
        for m in range(0, len(title)):
            worksheet2.write(0, m, title[m], bold_title)
            worksheet3.write(0, m, title[m], bold_title)
            # worksheet4.write(0, m, title[m], bold_title)


        # if msg1:
        #     for k in range(0, len(list_msg1)):
        #         for n in range(0, len(list_msg1[k])):
        #             worksheet2.write(k + 1, n, list(list_msg1[k])[n])

        if msg2:
            for i in range(0, len(list_msg2)):
                for j in range(0, len(list_msg2[i])):
                    worksheet2.write(i + 1, j, list_msg2[i][j])

        if msg3:
            for p in range(0, len(list_msg3)):
                for q in range(0, len(list_msg3[p])):
                    worksheet3.write(p + 1, q, list(list_msg3[p])[q])

        workbook.close()
        print('import report end')
        return workbookname, len(list_msg2), len(list_msg3)

    def second_dup(self, list1):
        second_dup_list = []
        for i in range(0, len(list1)):
            buffer_list = list1[i]
            for j in range(0, len(list1)):
                if buffer_list[8] == list1[j][8] and (i != j):
                    comp_result = difflib.SequenceMatcher(None, str(list1[j][-1]), buffer_list[-1]).quick_ratio()
                    comp_result2 = difflib.SequenceMatcher(None, str(list1[j][-1]), buffer_list[-1]).real_quick_ratio()
                    if comp_result > 0.8 and comp_result2 > 0.8:
                        list1[j] = '**********'
        # list1.remove('**********')
        for k in range(0, len(list1)):
            if list1[k] == '**********':
                pass
            else:
                second_dup_list.append(list1[k])
        # print(second_dup_list)
        return second_dup_list

    def split_list(self, str1):
        r = r'\'\t'
        strlist = str(str1).rstrip()
        count = str(strlist).count(' ')
        print('count=========', count)
        if count >= 5:
            (date, split1) = strlist.split(' ', 1)
            (time, split2) = split1.strip().split(' ', 1)
            (pid, split3) = split2.strip().split(' ', 1)
            (tid, split4) = split3.strip().split(' ', 1)
            (logver, split5) = split4.strip().split(' ', 1)
            (tag, split6) = split5.strip().split(':', 1)
            re_date = date.strip('b\'\"')
            ex_info = split6.rstrip(r)
            # print('+++++++++++++++++++++++', ex_info)
            return (re_date, time, pid, tid, logver, tag, ex_info)

    def get_DevicesID_dirs(self, path):
        print('get_DevicesID')
        dirs_list = []
        for dirs in os.listdir(path):
            dirs_list.append(dirs)
        print('dirs_list=========', dirs_list)
        return dirs_list

    def parse_all_log(self, path):
        files_list = []
        for root, dirs, files in os.walk(path):
            for i in range(0, len(files)):
                root_files = '%s\%s' % (root, files[i])
                root_files_buffer = str(root_files)
                # print('files======', root_files_buffer)
                files_list.append(root_files)
        # print('files list ========', files_list)
        return files_list


    def select_path(self):
        dir_path = filedialog.askdirectory()
        self.lineEdit_path.setText(dir_path)


def check_env(self):
    try:
        conn = sqlite3.connect('//192.168.3.88/sql/AutoToolCheck/ToolCheckEnv.db')
        print(conn)
        conn.close()
        return True
    except:
        tk.messagebox.showerror('ERROR', '无法运行该工具，该工具仅限内部使用,\n请确定已经正确连接内部服务器！')
        return True

if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)  # 首先必须实例化QApplication类，作为GUI主程序入口
    if check_env(self=None):
        MainWindow = MainWindow()  # 实例化MainWindow类，创建自带menu的窗体类型QMainWindow
        MainWindow.setFixedSize(574, 270)  # 禁止改变窗口大小
        MainWindow.show()  # 显示窗体
        sys.exit(app.exec_())  # 当来自操作系统的分发事件指派调用窗口时，
    else:
        sys.exit(app.exec_())
